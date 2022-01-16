import cv2
from tkinter import *
from tkinter import ttk
from tkinter import messagebox
from PIL import Image, ImageTk
from datetime  import date
from datetime import datetime
import os
from openpyxl import load_workbook, Workbook
import time
from pathlib import Path

#กำหนดค่าเริ่มต้นให้เป็นศูนย์
Incline_cnt = 0
Chip1_cnt = 0
Chip2_cnt = 0
Check_cnt = 0
Scratch_cnt =0
Scuff_cnt = 0
Stone_cnt = 0
Blister_cnt = 0
Airline_cnt = 0
Knot_cnt = 0
Other_cnt = 0
totalNG_cnt = 0
FG = 0

cap = 0 #กำหนดค่าตัวแปรเริ่มต้นสำหรับเปิดกล้องเท่ากับศูนย์

#clear ข้อมูลให้เป็นศูนย์และช่องว่างหลังจากกด save 
def Clear_alldata():
    global Incline_cnt,Chip1_cnt,Chip2_cnt,Check_cnt,Scratch_cnt, Scuff_cnt, Stone_cnt, Blister_cnt, Airline_cnt, Knot_cnt, Other_cnt, totalNG_cnt, FG

    Incline_cnt = 0
    Chip1_cnt = 0
    Chip2_cnt = 0
    Check_cnt = 0
    Scratch_cnt =0
    Scuff_cnt = 0
    Stone_cnt = 0
    Blister_cnt = 0
    Airline_cnt = 0
    Knot_cnt = 0
    Other_cnt = 0
    totalNG_cnt = 0
    FG = 0

    V_itemname.set(' ')
    V_RawmatLot.set('')
    V_CuttingLot.set('')
    V_input.set('')
    V_FG.set(0)
    V_totalNG.set(0)
    V_percent.set(0)
    V_Incline.set(0)
    V_Chip1.set(0)
    V_Chip2.set(0)
    V_check.set(0)
    V_scr.set(0)
    V_scf.set(0)
    V_stone.set(0)
    V_BL.set(0)
    V_AL.set(0)
    V_knot.set(0)
    V_Other.set(0)
    V_name.set('')
    V_remark.set('')

    itemname.focus() # ให้ curser ไปอยู่ที่ช่องกรอก itemname หลังจากกด save&clear 

def SavetoExcel():

    #set up file name, excel file must be setting in folder
    workbook_name = Path(r'D:\Tubecutting\data.xlsx')
    
    if workbook_name.exists():

        wb = load_workbook(workbook_name) 
        sheet = wb.active

        dt = datetime.now() #แสดงวันเวลาปัจจุบัน
        dt = datetime.strftime(dt,'%Y-%m-%d %H:%M')

        #set up header name

        sheet['A1'].value = 'Item Name'
        sheet['B1'].value = 'RawMat Lot No.'
        sheet['C1'].value = 'Cutting Lot No.'
        sheet['D1'].value = 'Input(pcs.)'
        sheet['E1'].value = 'FG (pcs.)'
        sheet['F1'].value = 'Total NG(pcs.)'
        sheet['G1'].value = 'FG (%)'
        sheet['H1'].value = 'Incline(pcs.)'
        sheet['I1'].value = 'Chip1(pcs.)'
        sheet['J1'].value = 'Chip2(pcs.)'
        sheet['K1'].value = 'Check(pcs.)'
        sheet['L1'].value = 'Scratch(pcs.)'
        sheet['M1'].value = 'Scuff(pcs.)'
        sheet['N1'].value = 'Stone(pcs.)'
        sheet['O1'].value = 'Blister(pcs.)'
        sheet['P1'].value = 'Airline(pcs.)'
        sheet['Q1'].value = 'Knot(pcs.)'
        sheet['R1'].value = 'Other(pcs.)'
        sheet['S1'].value = 'Inspector Name'
        sheet['T1'].value = 'Time Record'
        sheet['U1'].value = 'Remark'
 
        #new data to write to excel
    
        sheet.cell(column = 1, row =sheet.max_row +1 , value = V_itemname.get()) # Maxrow+1 หมายถึงการขึ้นบรรทัดใหม่
        sheet.cell(column = 2, row =sheet.max_row, value = V_RawmatLot.get())
        sheet.cell(column = 3, row =sheet.max_row , value = V_CuttingLot.get())
        sheet.cell(column = 4, row =sheet.max_row , value = V_input.get())
        sheet.cell(column = 5, row =sheet.max_row , value = V_FG.get())
        sheet.cell(column = 6, row =sheet.max_row , value = V_totalNG.get())
        sheet.cell(column = 7, row =sheet.max_row , value = float(V_percent.get()))
        sheet.cell(column = 8, row =sheet.max_row , value = V_Incline.get())
        sheet.cell(column = 9, row =sheet.max_row , value = V_Chip1.get())
        sheet.cell(column = 10, row =sheet.max_row , value = V_Chip2.get())
        sheet.cell(column = 11, row =sheet.max_row , value = V_check.get())
        sheet.cell(column = 12, row =sheet.max_row , value = V_scr.get())
        sheet.cell(column = 13, row =sheet.max_row , value = V_scf.get())
        sheet.cell(column = 14, row =sheet.max_row , value = V_stone.get())
        sheet.cell(column = 15, row =sheet.max_row , value = V_BL.get())
        sheet.cell(column = 16, row =sheet.max_row , value = V_AL.get())
        sheet.cell(column = 17, row =sheet.max_row , value = V_knot.get())
        sheet.cell(column = 18, row =sheet.max_row , value = V_Other.get())
        sheet.cell(column = 19, row =sheet.max_row , value = V_name.get())
        sheet.cell(column = 20, row =sheet.max_row , value = dt)
        sheet.cell(column = 21, row =sheet.max_row , value = V_remark.get())

        wb.save(workbook_name)
    else:
        messagebox.showwarning('loss file','no file name - data.xlsx')

    Clear_alldata()

def IncPlus_click():
    global Incline_cnt
    global totalNG_cnt

    try: 
        
        if V_input.get() == 0: # ะแสดงข้อความเตือนเมื่อข้อมูลช่อง Input เป็นศูนย์
            messagebox.showwarning('Warning', 'ช่อง Input ต้องเไม่เป็นศูนย์ครับ!')

        else:
            Incline_cnt += 1  #  บวกเพิ่มครั้งละ 1 เมื่อคลิ๊กปุ่ม 
            V_Incline.set(Incline_cnt) #เก็บค่าในตัวแปร V_incline

            totalNG_cnt +=1  #  บวกเพิ่ม Total NG ครั้งละ 1 เมื่อคลิ๊กปุ่ม 
            V_totalNG.set(totalNG_cnt) # เก็บค่าในตัวแปร V_totalNG

            FG = (V_input.get() - totalNG_cnt) # คำนวณหาจำนวน FG

            if FG <0:
                messagebox.showwarning('Warning','จำนวนไม่บาลานซ์ครับ')

            elif totalNG_cnt <0:
                messagebox.showwarning('Warning','จำนวนต้องไม่่ติดลบครับ')             

            else:    
                V_FG.set(int(FG)) # เก็บค่า  FG ลงในตัวแปร V_FG
                FG_percent = (FG/V_input.get())*100
                FG_percent = round(FG_percent,1) #แสดงทศนิยม 1 ตำแหน่ง 
                V_percent.set('{:.1f}'.format(FG_percent))

    except:
        messagebox.showwarning('Warning', 'กรุณากรอกข้อมูลให้ถูกต้องครบถ้วนครับ!')

def IncMinus_click():
    global Incline_cnt 
    global totalNG_cnt

    try:
        if V_input.get() == 0: # ะแสดงข้อความเตือนเมื่อข้อมูลช่อง Input เป็นศูนย์
            messagebox.showwarning('Warning', 'ช่อง Input ต้องเไม่เป็นศูนย์ครับ!')

        else:

            Incline_cnt -= 1  #  ลดครั้งละ 1 เมื่อคลิ๊กปุ่ม 
            V_Incline.set(Incline_cnt) #เก็บค่าในตัวแปร V_incline

            totalNG_cnt -=1  #  ลด Total NG ครั้งละ 1 เมื่อคลิ๊กปุ่ม 
            V_totalNG.set(totalNG_cnt) # เก็บค่าในตัวแปร V_totalNG

            FG = (V_input.get() - totalNG_cnt) # คำนวณหาจำนวน FG

            if FG <0:
                messagebox.showwarning('Warning','จำนวนไม่บาลานซ์ครับ')

            elif totalNG_cnt <0:
                messagebox.showwarning('Warning','จำนวนต้องไม่่ติดลบครับ')

            elif Incline_cnt <0:
                messagebox.showwarning('Warning','จำนวนต้องไม่่ติดลบครับ')   

            else:    
                V_FG.set(int(FG)) # เก็บค่า  FG ลงในตัวแปร V_FG
                FG_percent = (FG/V_input.get())*100
                FG_percent = round(FG_percent,1) #แสดงทศนิยม 1 ตำแหน่ง 
                V_percent.set('{:.1f}'.format(FG_percent))

    except:
        messagebox.showwarning('Warning', 'กรุณากรอกข้อมูลให้ถูกต้องครบถ้วนครับ!')

def Chip1Plus_click():
    global Chip1_cnt
    global totalNG_cnt

    try:
        if V_input.get() == 0: # ะแสดงข้อความเตือนเมื่อข้อมูลช่อง Input เป็นศูนย์
            messagebox.showwarning('Warning', 'ช่อง Input ต้องเไม่เป็นศูนย์ครับ!')

        else:
            Chip1_cnt += 1  #  บวกเพิ่มครั้งละ 1 เมื่อคลิ๊กปุ่ม 
            V_Chip1.set(Chip1_cnt) #เก็บค่าในตัวแปร V_chip1

            totalNG_cnt +=1  #  บวกเพิ่ม Total NG ครั้งละ 1 เมื่อคลิ๊กปุ่ม 
            V_totalNG.set(totalNG_cnt) # เก็บค่าในตัวแปร V_totalNG

            FG = (V_input.get() - totalNG_cnt) # คำนวณหาจำนวน FG

            if FG <0:
                messagebox.showwarning('Warning','จำนวนไม่บาลานซ์ครับ')

            elif totalNG_cnt <0:
                messagebox.showwarning('Warning','จำนวนต้องไม่่ติดลบครับ')             

            else:    
                V_FG.set(int(FG)) # เก็บค่า  FG ลงในตัวแปร V_FG
                FG_percent = (FG/V_input.get())*100
                FG_percent = round(FG_percent,1) #แสดงทศนิยม 1 ตำแหน่ง 
                V_percent.set('{:.1f}'.format(FG_percent))

    except:
        messagebox.showwarning('Warning', 'กรุณากรอกข้อมูลให้ถูกต้องครบถ้วนครับ!')

def Chip1Minus_click():
    global Chip1_cnt
    global totalNG_cnt

    try:
        if V_input.get() == 0: # ะแสดงข้อความเตือนเมื่อข้อมูลช่อง Input เป็นศูนย์
            messagebox.showwarning('Warning', 'ช่อง Input ต้องเไม่เป็นศูนย์ครับ!')

        else:

            Chip1_cnt -= 1  #  ลดครั้งละ 1 เมื่อคลิ๊กปุ่ม 
            V_Chip1.set(Chip1_cnt) #เก็บค่าในตัวแปร V_chip1

            totalNG_cnt -=1  #  ลด Total NG ครั้งละ 1 เมื่อคลิ๊กปุ่ม 
            V_totalNG.set(totalNG_cnt) # เก็บค่าในตัวแปร V_totalNG

            FG = (V_input.get() - totalNG_cnt) # คำนวณหาจำนวน FG

            if FG <0:
                messagebox.showwarning('Warning','จำนวนไม่บาลานซ์ครับ')

            elif totalNG_cnt <0:
                messagebox.showwarning('Warning','จำนวนต้องไม่่ติดลบครับ')

            elif Chip1_cnt <0:
                messagebox.showwarning('Warning','จำนวนต้องไม่่ติดลบครับ')   

            else:    
                V_FG.set(int(FG)) # เก็บค่า  FG ลงในตัวแปร V_FG
                FG_percent = (FG/V_input.get())*100
                FG_percent = round(FG_percent,1) #แสดงทศนิยม 1 ตำแหน่ง 
                V_percent.set('{:.1f}'.format(FG_percent))

    except:
        messagebox.showwarning('Warning', 'กรุณากรอกข้อมูลให้ถูกต้องครบถ้วนครับ!')

def Chip2Plus_click():
    global Chip2_cnt
    global totalNG_cnt

    try:
        if V_input.get() == 0: # ะแสดงข้อความเตือนเมื่อข้อมูลช่อง Input เป็นศูนย์
            messagebox.showwarning('Warning', 'ช่อง Input ต้องเไม่เป็นศูนย์ครับ!')

        else:

            Chip2_cnt += 1  #  ลดครั้งละ 1 เมื่อคลิ๊กปุ่ม 
            V_Chip2.set(Chip2_cnt) #เก็บค่าในตัวแปร V_Chip2

            totalNG_cnt +=1  #  ลด Total NG ครั้งละ 1 เมื่อคลิ๊กปุ่ม 
            V_totalNG.set(totalNG_cnt) # เก็บค่าในตัวแปร V_totalNG

            FG = (V_input.get() - totalNG_cnt) # คำนวณหาจำนวน FG

            if FG <0:
                messagebox.showwarning('Warning','จำนวนไม่บาลานซ์ครับ')

            elif totalNG_cnt <0:
                messagebox.showwarning('Warning','จำนวนต้องไม่่ติดลบครับ')

            elif Chip2_cnt <0:
                messagebox.showwarning('Warning','จำนวนต้องไม่่ติดลบครับ')   

            else:    
                V_FG.set(int(FG)) # เก็บค่า  FG ลงในตัวแปร V_FG
                FG_percent = (FG/V_input.get())*100
                FG_percent = round(FG_percent,1) #แสดงทศนิยม 1 ตำแหน่ง 
                V_percent.set('{:.1f}'.format(FG_percent))

    except:
        messagebox.showwarning('Warning', 'กรุณากรอกข้อมูลให้ถูกต้องครบถ้วนครับ!')


def Chip2Minus_click():
    global Chip2_cnt
    global totalNG_cnt

    try:
        if V_input.get() == 0: # ะแสดงข้อความเตือนเมื่อข้อมูลช่อง Input เป็นศูนย์
            messagebox.showwarning('Warning', 'ช่อง Input ต้องเไม่เป็นศูนย์ครับ!')

        else:

            Chip2_cnt -= 1  #  ลดครั้งละ 1 เมื่อคลิ๊กปุ่ม 
            V_Chip2.set(Chip2_cnt) #เก็บค่าในตัวแปร V_Chip2

            totalNG_cnt -=1  #  ลด Total NG ครั้งละ 1 เมื่อคลิ๊กปุ่ม 
            V_totalNG.set(totalNG_cnt) # เก็บค่าในตัวแปร V_totalNG

            FG = (V_input.get() - totalNG_cnt) # คำนวณหาจำนวน FG

            if FG <0:
                messagebox.showwarning('Warning','จำนวนไม่บาลานซ์ครับ')

            elif totalNG_cnt <0:
                messagebox.showwarning('Warning','จำนวนต้องไม่่ติดลบครับ')

            elif Chip2_cnt <0:
                messagebox.showwarning('Warning','จำนวนต้องไม่่ติดลบครับ')   

            else:    
                V_FG.set(int(FG)) # เก็บค่า  FG ลงในตัวแปร V_FG
                FG_percent = (FG/V_input.get())*100
                FG_percent = round(FG_percent,1) #แสดงทศนิยม 1 ตำแหน่ง 
                V_percent.set('{:.1f}'.format(FG_percent))

    except:
        messagebox.showwarning('Warning', 'กรุณากรอกข้อมูลให้ถูกต้องครบถ้วนครับ!')

def CheckPlus_click():
    global Check_cnt
    global totalNG_cnt

    try:
        if V_input.get() == 0: # ะแสดงข้อความเตือนเมื่อข้อมูลช่อง Input เป็นศูนย์
            messagebox.showwarning('Warning', 'ช่อง Input ต้องเไม่เป็นศูนย์ครับ!')

        else:

            Check_cnt += 1  #  ลดครั้งละ 1 เมื่อคลิ๊กปุ่ม 
            V_check.set(Check_cnt) #เก็บค่าในตัวแปร V_check

            totalNG_cnt +=1  #  ลด Total NG ครั้งละ 1 เมื่อคลิ๊กปุ่ม 
            V_totalNG.set(totalNG_cnt) # เก็บค่าในตัวแปร V_totalNG

            FG = (V_input.get() - totalNG_cnt) # คำนวณหาจำนวน FG

            if FG <0:
                messagebox.showwarning('Warning','จำนวนไม่บาลานซ์ครับ')

            elif totalNG_cnt <0:
                messagebox.showwarning('Warning','จำนวนต้องไม่่ติดลบครับ')

            elif Check_cnt <0:
                messagebox.showwarning('Warning','จำนวนต้องไม่่ติดลบครับ')   

            else:    
                V_FG.set(int(FG)) # เก็บค่า  FG ลงในตัวแปร V_FG
                FG_percent = (FG/V_input.get())*100
                FG_percent = round(FG_percent,1) #แสดงทศนิยม 1 ตำแหน่ง 
                V_percent.set('{:.1f}'.format(FG_percent))

    except:
        messagebox.showwarning('Warning', 'กรุณากรอกข้อมูลให้ถูกต้องครบถ้วนครับ!')

def CheckMinus_click():
    global Check_cnt
    global totalNG_cnt

    try:
        if V_input.get() == 0: # ะแสดงข้อความเตือนเมื่อข้อมูลช่อง Input เป็นศูนย์
            messagebox.showwarning('Warning', 'ช่อง Input ต้องเไม่เป็นศูนย์ครับ!')

        else:

            Check_cnt -= 1  #  ลดครั้งละ 1 เมื่อคลิ๊กปุ่ม 
            V_check.set(Check_cnt) #เก็บค่าในตัวแปร V_check

            totalNG_cnt -=1  #  ลด Total NG ครั้งละ 1 เมื่อคลิ๊กปุ่ม 
            V_totalNG.set(totalNG_cnt) # เก็บค่าในตัวแปร V_totalNG

            FG = (V_input.get() - totalNG_cnt) # คำนวณหาจำนวน FG

            if FG <0:
                messagebox.showwarning('Warning','จำนวนไม่บาลานซ์ครับ')

            elif totalNG_cnt <0:
                messagebox.showwarning('Warning','จำนวนต้องไม่่ติดลบครับ')

            elif Check_cnt <0:
                messagebox.showwarning('Warning','จำนวนต้องไม่่ติดลบครับ')   

            else:    
                V_FG.set(int(FG)) # เก็บค่า  FG ลงในตัวแปร V_FG
                FG_percent = (FG/V_input.get())*100
                FG_percent = round(FG_percent,1) #แสดงทศนิยม 1 ตำแหน่ง 
                V_percent.set('{:.1f}'.format(FG_percent))

    except:
        messagebox.showwarning('Warning', 'กรุณากรอกข้อมูลให้ถูกต้องครบถ้วนครับ!')

def SCRPlus_click():
    global Scratch_cnt
    global totalNG_cnt

    try:
        if V_input.get() == 0: # ะแสดงข้อความเตือนเมื่อข้อมูลช่อง Input เป็นศูนย์
            messagebox.showwarning('Warning', 'ช่อง Input ต้องเไม่เป็นศูนย์ครับ!')

        else:

            Scratch_cnt += 1  #  ลดครั้งละ 1 เมื่อคลิ๊กปุ่ม 
            V_scr.set(Scratch_cnt) #เก็บค่าในตัวแปร V_scr

            totalNG_cnt +=1  #  ลด Total NG ครั้งละ 1 เมื่อคลิ๊กปุ่ม 
            V_totalNG.set(totalNG_cnt) # เก็บค่าในตัวแปร V_totalNG

            FG = (V_input.get() - totalNG_cnt) # คำนวณหาจำนวน FG

            if FG <0:
                messagebox.showwarning('Warning','จำนวนไม่บาลานซ์ครับ')

            elif totalNG_cnt <0:
                messagebox.showwarning('Warning','จำนวนต้องไม่่ติดลบครับ')

            elif Scratch_cnt <0:
                messagebox.showwarning('Warning','จำนวนต้องไม่่ติดลบครับ')   

            else:    
                V_FG.set(int(FG)) # เก็บค่า  FG ลงในตัวแปร V_FG
                FG_percent = (FG/V_input.get())*100
                FG_percent = round(FG_percent,1) #แสดงทศนิยม 1 ตำแหน่ง 
                V_percent.set('{:.1f}'.format(FG_percent))

    except:
        messagebox.showwarning('Warning', 'กรุณากรอกข้อมูลให้ถูกต้องครบถ้วนครับ!')

def SCRMinus_click():
    global Scratch_cnt
    global totalNG_cnt

    try:
        if V_input.get() == 0: # ะแสดงข้อความเตือนเมื่อข้อมูลช่อง Input เป็นศูนย์
            messagebox.showwarning('Warning', 'ช่อง Input ต้องเไม่เป็นศูนย์ครับ!')

        else:

            Scratch_cnt -= 1  #  ลดครั้งละ 1 เมื่อคลิ๊กปุ่ม 
            V_scr.set(Scratch_cnt) #เก็บค่าในตัวแปร V_scr

            totalNG_cnt -=1  #  ลด Total NG ครั้งละ 1 เมื่อคลิ๊กปุ่ม 
            V_totalNG.set(totalNG_cnt) # เก็บค่าในตัวแปร V_totalNG

            FG = (V_input.get() - totalNG_cnt) # คำนวณหาจำนวน FG

            if FG <0:
                messagebox.showwarning('Warning','จำนวนไม่บาลานซ์ครับ')

            elif totalNG_cnt <0:
                messagebox.showwarning('Warning','จำนวนต้องไม่่ติดลบครับ')

            elif Scratch_cnt <0:
                messagebox.showwarning('Warning','จำนวนต้องไม่่ติดลบครับ')   

            else:    
                V_FG.set(int(FG)) # เก็บค่า  FG ลงในตัวแปร V_FG
                FG_percent = (FG/V_input.get())*100
                FG_percent = round(FG_percent,1) #แสดงทศนิยม 1 ตำแหน่ง 
                V_percent.set('{:.1f}'.format(FG_percent))

    except:
        messagebox.showwarning('Warning', 'กรุณากรอกข้อมูลให้ถูกต้องครบถ้วนครับ!')

def SCFPlus_click():
    global Scuff_cnt
    global totalNG_cnt

    try:
        if V_input.get() == 0: # ะแสดงข้อความเตือนเมื่อข้อมูลช่อง Input เป็นศูนย์
            messagebox.showwarning('Warning', 'ช่อง Input ต้องเไม่เป็นศูนย์ครับ!')

        else:

            Scuff_cnt += 1  #  ลดครั้งละ 1 เมื่อคลิ๊กปุ่ม 
            V_scf.set(Scuff_cnt) #เก็บค่าในตัวแปร V_scf

            totalNG_cnt +=1  #  ลด Total NG ครั้งละ 1 เมื่อคลิ๊กปุ่ม 
            V_totalNG.set(totalNG_cnt) # เก็บค่าในตัวแปร V_totalNG

            FG = (V_input.get() - totalNG_cnt) # คำนวณหาจำนวน FG

            if FG <0:
                messagebox.showwarning('Warning','จำนวนไม่บาลานซ์ครับ')

            elif totalNG_cnt <0:
                messagebox.showwarning('Warning','จำนวนต้องไม่่ติดลบครับ')

            elif Scuff_cnt <0:
                messagebox.showwarning('Warning','จำนวนต้องไม่่ติดลบครับ')   

            else:    
                V_FG.set(int(FG)) # เก็บค่า  FG ลงในตัวแปร V_FG
                FG_percent = (FG/V_input.get())*100
                FG_percent = round(FG_percent,1) #แสดงทศนิยม 1 ตำแหน่ง 
                V_percent.set('{:.1f}'.format(FG_percent))

    except:
        messagebox.showwarning('Warning', 'กรุณากรอกข้อมูลให้ถูกต้องครบถ้วนครับ!')
                
def SCFMinus_click():
    global Scuff_cnt
    global totalNG_cnt

    try:
        if V_input.get() == 0: # ะแสดงข้อความเตือนเมื่อข้อมูลช่อง Input เป็นศูนย์
            messagebox.showwarning('Warning', 'ช่อง Input ต้องเไม่เป็นศูนย์ครับ!')

        else:

            Scuff_cnt -= 1  #  ลดครั้งละ 1 เมื่อคลิ๊กปุ่ม 
            V_scf.set(Scuff_cnt) #เก็บค่าในตัวแปร V_scf

            totalNG_cnt -=1  #  ลด Total NG ครั้งละ 1 เมื่อคลิ๊กปุ่ม 
            V_totalNG.set(totalNG_cnt) # เก็บค่าในตัวแปร V_totalNG

            FG = (V_input.get() - totalNG_cnt) # คำนวณหาจำนวน FG

            if FG <0:
                messagebox.showwarning('Warning','จำนวนไม่บาลานซ์ครับ')

            elif totalNG_cnt <0:
                messagebox.showwarning('Warning','จำนวนต้องไม่่ติดลบครับ')

            elif Scuff_cnt <0:
                messagebox.showwarning('Warning','จำนวนต้องไม่่ติดลบครับ')   

            else:    
                V_FG.set(int(FG)) # เก็บค่า  FG ลงในตัวแปร V_FG
                FG_percent = (FG/V_input.get())*100
                FG_percent = round(FG_percent,1) #แสดงทศนิยม 1 ตำแหน่ง 
                V_percent.set('{:.1f}'.format(FG_percent))

    except:
        messagebox.showwarning('Warning', 'กรุณากรอกข้อมูลให้ถูกต้องครบถ้วนครับ!')

def STPlus_click():
    global Stone_cnt
    global totalNG_cnt

    try:
        if V_input.get() == 0: # ะแสดงข้อความเตือนเมื่อข้อมูลช่อง Input เป็นศูนย์
            messagebox.showwarning('Warning', 'ช่อง Input ต้องเไม่เป็นศูนย์ครับ!')

        else:

            Stone_cnt += 1  #  ลดครั้งละ 1 เมื่อคลิ๊กปุ่ม 
            V_stone.set(Stone_cnt) #เก็บค่าในตัวแปร V_stone

            totalNG_cnt +=1  #  ลด Total NG ครั้งละ 1 เมื่อคลิ๊กปุ่ม 
            V_totalNG.set(totalNG_cnt) # เก็บค่าในตัวแปร V_totalNG

            FG = (V_input.get() - totalNG_cnt) # คำนวณหาจำนวน FG

            if FG <0:
                messagebox.showwarning('Warning','จำนวนไม่บาลานซ์ครับ')

            elif totalNG_cnt <0:
                messagebox.showwarning('Warning','จำนวนต้องไม่่ติดลบครับ')

            elif Stone_cnt <0:
                messagebox.showwarning('Warning','จำนวนต้องไม่่ติดลบครับ')   

            else:    
                V_FG.set(int(FG)) # เก็บค่า  FG ลงในตัวแปร V_FG
                FG_percent = (FG/V_input.get())*100
                FG_percent = round(FG_percent,1) #แสดงทศนิยม 1 ตำแหน่ง 
                V_percent.set('{:.1f}'.format(FG_percent))

    except:
        messagebox.showwarning('Warning', 'กรุณากรอกข้อมูลให้ถูกต้องครบถ้วนครับ!')


def STMinus_click():
    global Stone_cnt
    global totalNG_cnt

    try:
        if V_input.get() == 0: # ะแสดงข้อความเตือนเมื่อข้อมูลช่อง Input เป็นศูนย์
            messagebox.showwarning('Warning', 'ช่อง Input ต้องเไม่เป็นศูนย์ครับ!')

        else:

            Stone_cnt -= 1  #  ลดครั้งละ 1 เมื่อคลิ๊กปุ่ม 
            V_stone.set(Stone_cnt) #เก็บค่าในตัวแปร V_stone

            totalNG_cnt -=1  #  ลด Total NG ครั้งละ 1 เมื่อคลิ๊กปุ่ม 
            V_totalNG.set(totalNG_cnt) # เก็บค่าในตัวแปร V_totalNG

            FG = (V_input.get() - totalNG_cnt) # คำนวณหาจำนวน FG

            if FG <0:
                messagebox.showwarning('Warning','จำนวนไม่บาลานซ์ครับ')

            elif totalNG_cnt <0:
                messagebox.showwarning('Warning','จำนวนต้องไม่่ติดลบครับ')

            elif Stone_cnt <0:
                messagebox.showwarning('Warning','จำนวนต้องไม่่ติดลบครับ')   

            else:    
                V_FG.set(int(FG)) # เก็บค่า  FG ลงในตัวแปร V_FG
                FG_percent = (FG/V_input.get())*100
                FG_percent = round(FG_percent,1) #แสดงทศนิยม 1 ตำแหน่ง 
                V_percent.set('{:.1f}'.format(FG_percent))

    except:
        messagebox.showwarning('Warning', 'กรุณากรอกข้อมูลให้ถูกต้องครบถ้วนครับ!')

def BLPlus_click():
    global Blister_cnt
    global totalNG_cnt

    try:
        if V_input.get() == 0: # ะแสดงข้อความเตือนเมื่อข้อมูลช่อง Input เป็นศูนย์
            messagebox.showwarning('Warning', 'ช่อง Input ต้องเไม่เป็นศูนย์ครับ!')

        else:

            Blister_cnt += 1  #  ลดครั้งละ 1 เมื่อคลิ๊กปุ่ม 
            V_BL.set(Blister_cnt) #เก็บค่าในตัวแปร V_BL

            totalNG_cnt +=1  #  ลด Total NG ครั้งละ 1 เมื่อคลิ๊กปุ่ม 
            V_totalNG.set(totalNG_cnt) # เก็บค่าในตัวแปร V_totalNG

            FG = (V_input.get() - totalNG_cnt) # คำนวณหาจำนวน FG

            if FG <0:
                messagebox.showwarning('Warning','จำนวนไม่บาลานซ์ครับ')

            elif totalNG_cnt <0:
                messagebox.showwarning('Warning','จำนวนต้องไม่่ติดลบครับ')

            elif Blister_cnt <0:
                messagebox.showwarning('Warning','จำนวนต้องไม่่ติดลบครับ')   

            else:    
                V_FG.set(int(FG)) # เก็บค่า  FG ลงในตัวแปร V_FG
                FG_percent = (FG/V_input.get())*100
                FG_percent = round(FG_percent,1) #แสดงทศนิยม 1 ตำแหน่ง 
                V_percent.set('{:.1f}'.format(FG_percent))

    except:
        messagebox.showwarning('Warning', 'กรุณากรอกข้อมูลให้ถูกต้องครบถ้วนครับ!')

def BLMinus_click():
    global Blister_cnt
    global totalNG_cnt

    try:
        if V_input.get() == 0: # ะแสดงข้อความเตือนเมื่อข้อมูลช่อง Input เป็นศูนย์
            messagebox.showwarning('Warning', 'ช่อง Input ต้องเไม่เป็นศูนย์ครับ!')

        else:

            Blister_cnt -= 1  #  ลดครั้งละ 1 เมื่อคลิ๊กปุ่ม 
            V_BL.set(Blister_cnt) #เก็บค่าในตัวแปร V_BL

            totalNG_cnt -=1  #  ลด Total NG ครั้งละ 1 เมื่อคลิ๊กปุ่ม 
            V_totalNG.set(totalNG_cnt) # เก็บค่าในตัวแปร V_totalNG

            FG = (V_input.get() - totalNG_cnt) # คำนวณหาจำนวน FG

            if FG <0:
                messagebox.showwarning('Warning','จำนวนไม่บาลานซ์ครับ')

            elif totalNG_cnt <0:
                messagebox.showwarning('Warning','จำนวนต้องไม่่ติดลบครับ')

            elif Blister_cnt <0:
                messagebox.showwarning('Warning','จำนวนต้องไม่่ติดลบครับ')   

            else:    
                V_FG.set(int(FG)) # เก็บค่า  FG ลงในตัวแปร V_FG
                FG_percent = (FG/V_input.get())*100
                FG_percent = round(FG_percent,1) #แสดงทศนิยม 1 ตำแหน่ง 
                V_percent.set('{:.1f}'.format(FG_percent))

    except:
        messagebox.showwarning('Warning', 'กรุณากรอกข้อมูลให้ถูกต้องครบถ้วนครับ!')

def ALPlus_click():
    global Airline_cnt
    global totalNG_cnt

    try:
        if V_input.get() == 0: # ะแสดงข้อความเตือนเมื่อข้อมูลช่อง Input เป็นศูนย์
            messagebox.showwarning('Warning', 'ช่อง Input ต้องเไม่เป็นศูนย์ครับ!')

        else:

            Airline_cnt += 1  #  ลดครั้งละ 1 เมื่อคลิ๊กปุ่ม 
            V_AL.set(Airline_cnt) #เก็บค่าในตัวแปร V_AL

            totalNG_cnt +=1  #  ลด Total NG ครั้งละ 1 เมื่อคลิ๊กปุ่ม 
            V_totalNG.set(totalNG_cnt) # เก็บค่าในตัวแปร V_totalNG

            FG = (V_input.get() - totalNG_cnt) # คำนวณหาจำนวน FG

            if FG <0:
                messagebox.showwarning('Warning','จำนวนไม่บาลานซ์ครับ')

            elif totalNG_cnt <0:
                messagebox.showwarning('Warning','จำนวนต้องไม่่ติดลบครับ')

            elif Airline_cnt <0:
                messagebox.showwarning('Warning','จำนวนต้องไม่่ติดลบครับ')   

            else:    
                V_FG.set(int(FG)) # เก็บค่า  FG ลงในตัวแปร V_FG
                FG_percent = (FG/V_input.get())*100
                FG_percent = round(FG_percent,1) #แสดงทศนิยม 1 ตำแหน่ง 
                V_percent.set('{:.1f}'.format(FG_percent))

    except:
        messagebox.showwarning('Warning', 'กรุณากรอกข้อมูลให้ถูกต้องครบถ้วนครับ!')

def ALMinus_click():
    global Airline_cnt
    global totalNG_cnt

    try:
        if V_input.get() == 0: # ะแสดงข้อความเตือนเมื่อข้อมูลช่อง Input เป็นศูนย์
            messagebox.showwarning('Warning', 'ช่อง Input ต้องเไม่เป็นศูนย์ครับ!')

        else:

            Airline_cnt -= 1  #  ลดครั้งละ 1 เมื่อคลิ๊กปุ่ม 
            V_AL.set(Airline_cnt) #เก็บค่าในตัวแปร V_AL

            totalNG_cnt -=1  #  ลด Total NG ครั้งละ 1 เมื่อคลิ๊กปุ่ม 
            V_totalNG.set(totalNG_cnt) # เก็บค่าในตัวแปร V_totalNG

            FG = (V_input.get() - totalNG_cnt) # คำนวณหาจำนวน FG

            if FG <0:
                messagebox.showwarning('Warning','จำนวนไม่บาลานซ์ครับ')

            elif totalNG_cnt <0:
                messagebox.showwarning('Warning','จำนวนต้องไม่่ติดลบครับ')

            elif Airline_cnt <0:
                messagebox.showwarning('Warning','จำนวนต้องไม่่ติดลบครับ')   

            else:    
                V_FG.set(int(FG)) # เก็บค่า  FG ลงในตัวแปร V_FG
                FG_percent = (FG/V_input.get())*100
                FG_percent = round(FG_percent,1) #แสดงทศนิยม 1 ตำแหน่ง 
                V_percent.set('{:.1f}'.format(FG_percent))

    except:
        messagebox.showwarning('Warning', 'กรุณากรอกข้อมูลให้ถูกต้องครบถ้วนครับ!')

def knotPlus_click():
    global Knot_cnt
    global totalNG_cnt

    try:
        if V_input.get() == 0: # ะแสดงข้อความเตือนเมื่อข้อมูลช่อง Input เป็นศูนย์
            messagebox.showwarning('Warning', 'ช่อง Input ต้องเไม่เป็นศูนย์ครับ!')

        else:

            Knot_cnt += 1  #  ลดครั้งละ 1 เมื่อคลิ๊กปุ่ม 
            V_knot.set(Knot_cnt) #เก็บค่าในตัวแปร V_knot

            totalNG_cnt +=1  #  ลด Total NG ครั้งละ 1 เมื่อคลิ๊กปุ่ม 
            V_totalNG.set(totalNG_cnt) # เก็บค่าในตัวแปร V_totalNG

            FG = (V_input.get() - totalNG_cnt) # คำนวณหาจำนวน FG

            if FG <0:
                messagebox.showwarning('Warning','จำนวนไม่บาลานซ์ครับ')

            elif totalNG_cnt <0:
                messagebox.showwarning('Warning','จำนวนต้องไม่่ติดลบครับ')

            elif Knot_cnt <0:
                messagebox.showwarning('Warning','จำนวนต้องไม่่ติดลบครับ')   

            else:    
                V_FG.set(int(FG)) # เก็บค่า  FG ลงในตัวแปร V_FG
                FG_percent = (FG/V_input.get())*100
                FG_percent = round(FG_percent,1) #แสดงทศนิยม 1 ตำแหน่ง 
                V_percent.set('{:.1f}'.format(FG_percent))

    except:
        messagebox.showwarning('Warning', 'กรุณากรอกข้อมูลให้ถูกต้องครบถ้วนครับ!')

def knotMinus_click():
    global Knot_cnt
    global totalNG_cnt

    try:
        if V_input.get() == 0: # ะแสดงข้อความเตือนเมื่อข้อมูลช่อง Input เป็นศูนย์
            messagebox.showwarning('Warning', 'ช่อง Input ต้องเไม่เป็นศูนย์ครับ!')

        else:

            Knot_cnt -= 1  #  ลดครั้งละ 1 เมื่อคลิ๊กปุ่ม 
            V_knot.set(Knot_cnt) #เก็บค่าในตัวแปร V_knot

            totalNG_cnt -=1  #  ลด Total NG ครั้งละ 1 เมื่อคลิ๊กปุ่ม 
            V_totalNG.set(totalNG_cnt) # เก็บค่าในตัวแปร V_totalNG

            FG = (V_input.get() - totalNG_cnt) # คำนวณหาจำนวน FG

            if FG <0:
                messagebox.showwarning('Warning','จำนวนไม่บาลานซ์ครับ')

            elif totalNG_cnt <0:
                messagebox.showwarning('Warning','จำนวนต้องไม่่ติดลบครับ')

            elif Knot_cnt <0:
                messagebox.showwarning('Warning','จำนวนต้องไม่่ติดลบครับ')   

            else:    
                V_FG.set(int(FG)) # เก็บค่า  FG ลงในตัวแปร V_FG
                FG_percent = (FG/V_input.get())*100
                FG_percent = round(FG_percent,1) #แสดงทศนิยม 1 ตำแหน่ง 
                V_percent.set('{:.1f}'.format(FG_percent))

    except:
        messagebox.showwarning('Warning', 'กรุณากรอกข้อมูลให้ถูกต้องครบถ้วนครับ!')

def OtherPlus_click():
    global Other_cnt
    global totalNG_cnt

    try:
        if V_input.get() == 0: # ะแสดงข้อความเตือนเมื่อข้อมูลช่อง Input เป็นศูนย์
            messagebox.showwarning('Warning', 'ช่อง Input ต้องเไม่เป็นศูนย์ครับ!')

        else:

            Other_cnt += 1  #  ลดครั้งละ 1 เมื่อคลิ๊กปุ่ม 
            V_Other.set(Other_cnt) #เก็บค่าในตัวแปร V_Other

            totalNG_cnt +=1  #  ลด Total NG ครั้งละ 1 เมื่อคลิ๊กปุ่ม 
            V_totalNG.set(totalNG_cnt) # เก็บค่าในตัวแปร V_totalNG

            FG = (V_input.get() - totalNG_cnt) # คำนวณหาจำนวน FG

            if FG <0:
                messagebox.showwarning('Warning','จำนวนไม่บาลานซ์ครับ')

            elif totalNG_cnt <0:
                messagebox.showwarning('Warning','จำนวนต้องไม่่ติดลบครับ')

            elif Other_cnt <0:
                messagebox.showwarning('Warning','จำนวนต้องไม่่ติดลบครับ')   

            else:    
                V_FG.set(int(FG)) # เก็บค่า  FG ลงในตัวแปร V_FG
                FG_percent = (FG/V_input.get())*100
                FG_percent = round(FG_percent,1) #แสดงทศนิยม 1 ตำแหน่ง 
                V_percent.set('{:.1f}'.format(FG_percent))

    except:
        messagebox.showwarning('Warning', 'กรุณากรอกข้อมูลให้ถูกต้องครบถ้วนครับ!')

def OtherMinus_click():
    global Other_cnt
    global totalNG_cnt

    try:
        if V_input.get() == 0: # ะแสดงข้อความเตือนเมื่อข้อมูลช่อง Input เป็นศูนย์
            messagebox.showwarning('Warning', 'ช่อง Input ต้องเไม่เป็นศูนย์ครับ!')

        else:

            Other_cnt -= 1  #  ลดครั้งละ 1 เมื่อคลิ๊กปุ่ม 
            V_Other.set(Other_cnt) #เก็บค่าในตัวแปร V_Other

            totalNG_cnt -=1  #  ลด Total NG ครั้งละ 1 เมื่อคลิ๊กปุ่ม 
            V_totalNG.set(totalNG_cnt) # เก็บค่าในตัวแปร V_totalNG

            FG = (V_input.get() - totalNG_cnt) # คำนวณหาจำนวน FG

            if FG <0:
                messagebox.showwarning('Warning','จำนวนไม่บาลานซ์ครับ')

            elif totalNG_cnt <0:
                messagebox.showwarning('Warning','จำนวนต้องไม่่ติดลบครับ')

            elif Other_cnt <0:
                messagebox.showwarning('Warning','จำนวนต้องไม่่ติดลบครับ')   

            else:    
                V_FG.set(int(FG)) # เก็บค่า  FG ลงในตัวแปร V_FG
                FG_percent = (FG/V_input.get())*100
                FG_percent = round(FG_percent,1) #แสดงทศนิยม 1 ตำแหน่ง 
                V_percent.set('{:.1f}'.format(FG_percent))

    except:
        messagebox.showwarning('Warning', 'กรุณากรอกข้อมูลให้ถูกต้องครบถ้วนครับ!')

def  showframe():
	ret,frame = cap.read()
	if ret ==True:

		frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB) #เปลี่ยนระบบสีจาก BGR ไปเป็น RGB
		frame = cv2.flip(frame,1) # กลับภาพ ซ้าย ขวา
		img = cv2.resize(frame, None, fx=2,fy=2, interpolation = cv2.INTER_CUBIC) # ขยายภาพ 2 เท่า

		im = Image.fromarray(img) # เปลี่ยนภาพจาก Video frame เป็น Array
	                                                                                                                                                                                                                                    
		imgTK = ImageTk.PhotoImage(image=im) 

		Label1.configure(image=imgTK) #แปะภาพลงบนลาเบล GUI
		Label1.image = imgTK
		Label1.place(x=5,y=5)
		Label1.after(10,showframe)

	else:
			Label1.image = 'No camera signal'
			cap.release()

def StartVideo():
    global cap 
    cap =cv2.VideoCapture(0, cv2.CAP_DSHOW)
    showframe()

def StopVideo():
	global cap
	cap.release()

def exit():

    GUI.quit()

GUI=Tk()
GUI.geometry('1150x700')  # width x height
GUI.title('Tube Inspection v1.0')
#GUI.option_add("*Font",'consolas 15')

font1 = ('consolas', 16, 'bold')
font2  = ('consolas', 10, 'bold')
font3  = ('Tahoma', 10)

#Camera frame
frame1 = Frame(GUI, width = 1000, height = 300, bg ='light blue')
frame1.pack(side = TOP, fill =X, ipadx=5, ipady=5 )

#Start/Stop/Exit frame
frame2 = Frame(GUI, width = 200, height = 100)
frame2.pack(side = BOTTOM)

#defect frame
frame3 = Frame(GUI, width = 1000, height = 320, bg ='pink')
frame3.pack(side =TOP)

#result frame
frame4 = Frame(GUI, width = 1000, height = 200)
frame4.pack(side =TOP)

Label1 = Label(frame1)
Label2 = Label(frame2)


#Start button

Start = Button(frame2, text = 'START', font= font1, width = 12, command = StartVideo )
Start.pack(side = LEFT, ipadx= 5)

#Stop button

Stop = Button(frame2, text = 'STOP', font= font1, width = 12,fg = 'red', command = StopVideo)
Stop.pack(side = LEFT, ipadx= 5)

Savedata = Button(frame2, text= 'Save&Clear', font= font1, command = SavetoExcel, width = 12)
Savedata.pack(side = LEFT, ipadx= 5)

Exit = Button(frame2, text= 'Exit Program', font= font1, fg = 'red', bg= 'orange', width = 12, command = GUI.quit)
Exit.pack(side = RIGHT, ipadx= 5)

#defect button
photo1 = PhotoImage(file= 'plus.png')
photo2 = PhotoImage(file = 'minus.png')

#Incline defect
V_Incline = IntVar()

Defect1_1 = Button(frame3, text ='Incline', command = IncPlus_click,font= font2, image= photo1, compound = BOTTOM, width = 50, pady=5)
Defect1_1.grid(row = 0, column = 0, ipadx = 5)

Label1_1 = Label(frame3, textvariable = V_Incline, width = 5, font= font2)
Label1_1.grid(row = 1, column=0)

Defect1_2 = Button(frame3, text ='Incline', command = IncMinus_click,font= font2, image= photo2, compound = BOTTOM, width = 50, pady=5)
Defect1_2.grid(row = 2, column = 0, ipadx = 5)


#Chip1 defect
V_Chip1 = IntVar()
Defect2_1 = Button(frame3, text ='Chip1', command = Chip1Plus_click, font= font2, image= photo1, compound = BOTTOM, width = 50, pady=5)
Defect2_1.grid(row = 0, column = 1, ipadx = 5)

Label2_1 = Label(frame3, width = 5, textvariable = V_Chip1)
Label2_1.grid(row = 1, column=1)

Defect2_2 = Button(frame3, text ='Chip1', command = Chip1Minus_click, font= font2, image= photo2, compound = BOTTOM, width = 50, pady=5)
Defect2_2.grid(row = 2, column = 1, ipadx = 5)


#Chip2 defect
V_Chip2 = IntVar()
Defect3_1 = Button(frame3, text ='Chip2', command = Chip2Plus_click, font= font2, image= photo1, compound = BOTTOM, width = 50, pady=5)
Defect3_1.grid(row = 0, column = 2, ipadx = 5)

Label3_1 = Label(frame3, width = 5, textvariable = V_Chip2)
Label3_1.grid(row = 1, column=2)

Defect3_2 = Button(frame3, text ='Chip2', command = Chip2Minus_click,  font= font2, image= photo2, compound = BOTTOM, width = 50, pady=5)
Defect3_2.grid(row = 2, column = 2, ipadx = 5)


#Check defect
V_check = IntVar()
Defect4_1 = Button(frame3, text ='Check',command = CheckPlus_click,  font= font2, image= photo1, compound = BOTTOM, width = 50, pady=5)
Defect4_1.grid(row = 0, column = 3, ipadx = 5)

Label4_1 = Label(frame3, width = 5, textvariable= V_check)
Label4_1.grid(row = 1, column=3)

Defect4_2 = Button(frame3, text ='Check',command = CheckMinus_click,  font= font2, image= photo2, compound = BOTTOM, width = 50, pady=5)
Defect4_2.grid(row = 2, column = 3, ipadx = 5)

# Scratch defect
V_scr = IntVar()
Defect5_1 = Button(frame3, text ='Scratch', command = SCRPlus_click, font= font2, image= photo1, compound = BOTTOM, width = 50, pady=5)
Defect5_1.grid(row = 0, column = 4, ipadx = 5)

Label5_1 = Label(frame3, width = 5, textvariable = V_scr)
Label5_1.grid(row = 1, column=4)

Defect5_2 = Button(frame3, text ='Scratch',command = SCRMinus_click, font= font2, image= photo2, compound = BOTTOM, width = 50, pady=5)
Defect5_2.grid(row = 2, column = 4, ipadx = 5)


#Scuff defect
V_scf = IntVar()
Defect6_1 = Button(frame3, text ='Scuff',command = SCFPlus_click ,font= font2, image= photo1, compound = BOTTOM, width = 50, pady=5)
Defect6_1.grid(row = 0, column = 5, ipadx = 5)

Label6_1 = Label(frame3, width = 5, textvariable = V_scf)
Label6_1.grid(row = 1, column=5)

Defect6_2 = Button(frame3, text ='Scuff', command = SCFMinus_click,font= font2, image= photo2, compound = BOTTOM, width = 50, pady=5)
Defect6_2.grid(row = 2, column = 5, ipadx = 5)


#Stone defect
V_stone = IntVar()
Defect7_1 = Button(frame3, text ='Stone', command = STPlus_click, font= font2, image= photo1, compound = BOTTOM, width = 50, pady=5)
Defect7_1.grid(row = 0, column = 6, ipadx = 5)

Label7_1 = Label(frame3, width = 5, textvariable = V_stone)
Label7_1.grid(row = 1, column=6)

Defect7_2 = Button(frame3, text ='Stone', command = STMinus_click,  font= font2, image= photo2, compound = BOTTOM, width = 50, pady=5)
Defect7_2.grid(row = 2, column = 6, ipadx = 5)


#Blister defect
V_BL =  IntVar()
Defect8_1 = Button(frame3, text ='Blister', command = BLPlus_click, font= font2, image= photo1, compound = BOTTOM, width = 50, pady=5)
Defect8_1.grid(row = 0, column = 7, ipadx = 5)

Label8_1 = Label(frame3, width = 5, textvariable = V_BL)
Label8_1.grid(row = 1, column=7)

Defect8_2 = Button(frame3, text ='Blister',command = BLMinus_click, font= font2, image= photo2, compound = BOTTOM, width = 50, pady=5)
Defect8_2.grid(row = 2, column = 7, ipadx = 5)


#Airline defect
V_AL = IntVar()
Defect9_1 = Button(frame3, text ='Airline',command = ALPlus_click, font= font2, image= photo1, compound = BOTTOM, width = 50, pady=5)
Defect9_1.grid(row = 0, column = 8, ipadx = 5)

Label9_1 = Label(frame3, width = 5, textvariable = V_AL)
Label9_1.grid(row = 1, column=8)

Defect9_2 = Button(frame3, text ='Airline', command = ALMinus_click, font= font2, image= photo2, compound = BOTTOM, width = 50, pady=5)
Defect9_2.grid(row = 2, column = 8, ipadx = 5)

#Knot defect
V_knot = IntVar()
Defect10_1 = Button(frame3, text ='Knot', command = knotPlus_click, font= font2, image= photo1, compound = BOTTOM, width = 50, pady=5)
Defect10_1.grid(row = 0, column = 9, ipadx = 5)

Label10_1 = Label(frame3, width = 5, textvariable = V_knot)
Label10_1.grid(row = 1, column=9)

Defect10_2 = Button(frame3, text ='Knot', command = knotMinus_click, font= font2, image= photo2, compound = BOTTOM, width = 50, pady=5)
Defect10_2.grid(row = 2, column = 9, ipadx = 5)

#Other defect
V_Other = IntVar()
Defect11_1 = Button(frame3, text ='Other', command = OtherPlus_click, font= font2, image= photo1, compound = BOTTOM, width = 50, pady=5)
Defect11_1.grid(row = 0, column = 10, ipadx = 5)

Label11_1 = Label(frame3, width = 5, textvariable = V_Other)
Label11_1.grid(row = 1, column=10)

Defect11_2 = Button(frame3, text ='Other', command = OtherMinus_click,  font= font2, image= photo2, compound = BOTTOM, width = 50, pady=5)
Defect11_2.grid(row = 2, column = 10, ipadx = 5)


#Entry Item name
V_itemname = StringVar()
itemnameLabel = Label(frame4, text = 'Item Name', width = 10, font = font1, fg = 'black').grid(row = 3, column = 0, ipady = 10)
itemname = Entry(frame4, width= 12, font = font1, textvariable = V_itemname)
itemname.grid(row = 4,column = 0,ipady = 5)

itemname.focus() # show input curser when Start button press

#Entry Raw material Lot No.
V_RawmatLot = StringVar()
RawmatLabel = Label(frame4, text = 'Raw mat\nLot No.', width = 10, font = font1, fg = 'black').grid(row = 3, column = 1, ipady = 10)
Rawmat = Entry(frame4, width= 12, font = font1, textvariable = V_RawmatLot)
Rawmat.grid(row = 4,column = 1, ipady = 5)

#Entry  Cutting Lot No.
V_CuttingLot = StringVar()
CuttingLotLabel = Label(frame4, text = 'Cutting\nLot No.', width = 10, font = font1, fg = 'black').grid(row = 3, column = 2, ipady = 10)
CuttingLot = Entry(frame4, width= 12, font = font1, textvariable = V_CuttingLot)
CuttingLot.grid(row = 4,column = 2, ipady = 5)

#Entry Operator name
V_name = StringVar()
Name = Label(frame4, text = 'Operator\nName', width = 10, font = font1, fg = 'black').grid(row = 3, column = 3, ipady = 10)
Name = Entry(frame4, width= 10, font = font1, textvariable = V_name)
Name.grid(row = 4,column = 3, ipady = 5)

#Entry Input Quantity
V_input = IntVar()
V_input.set(' ')
LabelInput = Label(frame4, text = 'INPUT\n(pcs.)', width = 10, font = font1, fg = 'black').grid(row = 3, column = 4, ipady = 10)
Input = Entry(frame4, width= 8, font = font1, textvariable = V_input)
Input.grid(row = 4,column = 4, ipady = 5)


#Total NG quantity
V_totalNG = IntVar()
LabelNG = Label(frame4, text = 'Total NG', width = 10, font = font1, fg = 'black')
LabelNG.grid(row = 3, column = 5, ipady = 10, ipadx =10)

NG_pcs = Label(frame4, textvariable = V_totalNG,width = 10, font = font1,  fg = 'red').grid(row = 4, column = 5, ipady = 10, ipadx =10)

#Finish goods quantity
V_FG = IntVar()
LabelFG = Label(frame4, text = 'FG', width = 10, font = font1, fg = 'black').grid(row = 3, column = 6, ipady = 10, ipadx =10)
FG_pcs = Label(frame4, width = 10, textvariable = V_FG,  font = font1,  fg = 'green').grid(row = 4, column = 6, ipady = 10, ipadx =10)

#%FG
V_percent = StringVar()
LabelFG_percent = Label(frame4, text = '%FG', width = 10, font = font1, fg = 'black').grid(row = 3, column = 7, ipady = 10, ipadx =10)
FG_percent = Label(frame4, width = 10, textvariable = V_percent,  font = font1,  fg = 'green').grid(row = 4, column = 7, ipady = 10, ipadx =10)


#Entry Remark 
V_remark = StringVar()
RemarkLabel = Label(frame4, text = 'Remark', width = 10, font = font1, fg = 'black').grid(row = 5, column = 0, ipady = 10)
Remark = Entry(frame4, width= 100, font = font3 , textvariable = V_remark)
Remark.grid(row = 5,columnspan = 7, ipady = 5)


GUI.mainloop()